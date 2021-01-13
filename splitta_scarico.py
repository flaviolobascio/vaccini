import os
import time
import pandas as pd
from datetime import date
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import regex


data_location = 'scaricamenti/'
dati_splittati = 'splittati/'

new_file_directory = 'splittati/'
today = date.today()
oggi = today.strftime('%Y%m%d')
df_zero = pd.read_csv('df_zero.csv', sep=';')
df_zeroi = df_zero.set_index("Fascia età")
#print(df_zeroi)

nome_file_scarico = 'scarico_20210112.csv'
data_scarico_str = regex.findall('[0-9]{8}', nome_file_scarico)[0]
data_scarico = datetime.datetime.strptime(data_scarico_str, "%Y%m%d")
print(data_scarico_str)
df_scarico = pd.read_csv(data_location + nome_file_scarico, sep='\t', encoding='UTF-16')
df_codici = pd.read_csv('codici_punti_vaccinali.csv', sep=';', dtype={0: 'str', 1: 'str'})
df_codici_rev = pd.read_csv('codici_punti_reverse.csv', sep=';', dtype={0: 'str', 1: 'str'})
df_codici_idx = df_codici_rev.set_index("Codice Struttura")
print(df_codici_idx)
df_scarico_codici = pd.merge(df_scarico, df_codici, how='left', on='Centro Vaccinale')
#print(df_scarico_codici)
codici_unici = df_scarico_codici['Codice Struttura'].unique()
print(codici_unici)

for codice_struttura in codici_unici:

    nome_file_pv = data_scarico_str + "_" + codice_struttura + ".xlsx"
    shutil.copy("template.xlsx", dati_splittati + nome_file_pv)
    df_filter = df_scarico_codici[df_scarico_codici["Codice Struttura"] == codice_struttura]
    df_filter_gp = df_filter.groupby(['Codice Struttura', 'Fascia età', 'Codice Nsis']).sum()
    df_filter_gp2 = df_filter_gp[['Maschi', 'Femmine', 'Operatori Sanitari e Sociosanitari',
                                  'Personale non sanitario', 'Ospiti strutture residenziali',
                                  'Altro', 'Prima', 'Seconda']].reset_index(['Codice Struttura', 'Codice Nsis'])
    df_filter_gp3 = df_filter_gp2[['Maschi', 'Femmine', 'Operatori Sanitari e Sociosanitari',
                                  'Personale non sanitario', 'Ospiti strutture residenziali',
                                  'Altro', 'Prima', 'Seconda']]
    #print(df_filter_gp3)

    df_finale = pd.concat([df_zeroi, df_filter_gp3]).groupby("Fascia età")['Maschi', 'Femmine', 'Operatori Sanitari e Sociosanitari',
                                  'Personale non sanitario', 'Ospiti strutture residenziali',
                                  'Altro', 'Prima', 'Seconda'].sum()
    #print(df_finale)


    wb = load_workbook(dati_splittati + nome_file_pv)
    writer = pd.ExcelWriter(dati_splittati + nome_file_pv, engine='openpyxl')
    writer.book = wb
    ws = wb['Scheda']

    rows = dataframe_to_rows(df_finale, index=False, header=False)
    #print(rows)

    for r_idx, row in enumerate(rows, 0):
        for c_idx, value in enumerate(row, 0):
             ws.cell(row=r_idx+9, column=c_idx+4, value=value)

    # ws['E5'] = '123456789'
    ws['B5'] = data_scarico
    ws['E5'] = codice_struttura
    centro_vaccinale = df_codici_idx.loc[codice_struttura].values[0]
    print(centro_vaccinale)
    print(type(centro_vaccinale))
    ws['F5'] = centro_vaccinale
    ws['C2'] = "Rilevazione giornaliera Sommministrazione Vaccini anti-SARS-CoV-2/COVID-19 per punto vaccinale"
    # ws['D10'] = 150
    # ws['D11'] = 456
    wb.save(dati_splittati + nome_file_pv)


from pycel.excelcompiler import ExcelCompiler

# file_list = []
# totale_vaccini = 0
# for file in os.listdir(dati_splittati):
#     file_list.append(file)
#     xc = ExcelCompiler(filename=(dati_splittati + file))
#     a = xc.evaluate('Scheda!C18')
#     print(file, a)
#     totale_vaccini = totale_vaccini + a
#
# print(file_list)
# print('Totale vaccini: ', totale_vaccini)

file_list = []
totale_vaccini = 0
with os.scandir(dati_splittati) as it:
    for entry in it:
        if not entry.name.startswith('.') and entry.is_file():
            file_list.append(entry.name)
            xc = ExcelCompiler(filename=(dati_splittati + entry.name))
            a = xc.evaluate('Scheda!C18')
            print(entry.name, a)
            totale_vaccini = totale_vaccini + a
print(file_list)
print('Totale vaccini: ', totale_vaccini)
